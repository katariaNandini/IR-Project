import pandas as pd
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import re

# Load your data from the CSV file
df = pd.read_csv("clustered_posts_with_lda.csv")

# Organize posts by cluster and get a single dominant subtopic
clustered_data = defaultdict(list)
for _, row in df.iterrows():
    cluster_id = row['LDA_Topic']
    subtopic = row['LDA_Topic_Name']
    post_id = row['Post ID']
    text = row['Text']
    clustered_data[cluster_id].append((post_id, text, subtopic))

# Determine single dominant subtopic for each cluster
cluster_subtopics = {}
for cluster_id, posts in clustered_data.items():
    # Tokenize all words in subtopics and count their frequencies
    word_counter = Counter()
    for post_id, text, subtopic in posts:
        words = re.findall(r'\w+', subtopic.lower())  # Tokenize and convert to lowercase
        word_counter.update(words)
    
    # Get the most common word(s) as the dominant subtopic
    most_common_words = word_counter.most_common(2)  # Adjust to 1 or 2 for single or bigram labels
    dominant_subtopic = " ".join([word[0] for word in most_common_words])  # Combine as phrase
    
    cluster_subtopics[cluster_id] = dominant_subtopic

# Create an Excel workbook and add a worksheet
wb = Workbook()
ws = wb.active
ws.title = "Clustered Posts"

# Set column headers
ws.append(["Cluster ID", "Subtopic", "Post ID", "Text"])
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4F81BD")
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill

# Add data for each cluster
for cluster_id, posts in clustered_data.items():
    # Insert cluster row with the dominant subtopic
    dominant_subtopic = cluster_subtopics[cluster_id]
    cluster_row = [f"Cluster {cluster_id}", dominant_subtopic, "", ""]
    ws.append(cluster_row)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")
    
    # Add each post in the cluster
    for post_id, text, _ in posts:
        ws.append(["", "", post_id, text])
        
        # Apply wrap text to the 'Text' column for readability
        ws[f"D{ws.max_row}"].alignment = Alignment(wrap_text=True)

# Adjust column widths
ws.column_dimensions["A"].width = 15
ws.column_dimensions["B"].width = 25
ws.column_dimensions["C"].width = 10
ws.column_dimensions["D"].width = 100  # Increase for longer post texts

# Save the workbook
output_file = "clustered_posts_output.xlsx"
wb.save(output_file)

print(f"Structured output saved as {output_file}")
